import time
import pygame  # Importing the audio playback library
import random
import openpyxl
import logging
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from urllib.parse import urlparse

# Setting up the error log file
logging.basicConfig(filename="scraper_errors.log", level=logging.ERROR,
                    format="%(asctime)s - %(levelname)s - %(message)s")

# Setting up the Selenium browser
def setup_driver(proxy=None):
    chrome_options = Options()
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_argument("start-maximized")
    # chrome_options.add_argument("--headless")  # To run the browser without a GUI

    # Rotating the User-Agent
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Firefox/89.0",
    ]
    chrome_options.add_argument(f"user-agent={random.choice(user_agents)}")

    if proxy:
        chrome_options.add_argument(f"--proxy-server={proxy}")

    service = Service("C:/Users/hassa/Desktop/python/chromedriver.exe")  # Path to ChromeDriver
    return webdriver.Chrome(service=service, options=chrome_options)

# Generating Google search queries
def generate_queries(keywords, location):
    base_query_instagram = "site:instagram.com"
    base_query_facebook = "site:facebook.com"
    return [f"{base_query_instagram} {keyword} {location}" for keyword in keywords] + \
           [f"{base_query_facebook} {keyword} {location}" for keyword in keywords]

# Parsing search results using CSS Selectors
def parse_results(page_source):
    soup = BeautifulSoup(page_source, "html.parser")
    links = soup.select("a[href*='instagram.com'], a[href*='facebook.com']")
    valid_links = [link['href'] for link in links if is_valid_link(link['href'])]
    return list(set(valid_links))  # Remove duplicates

# Verifying the validity of links (Instagram or Facebook)
def is_valid_link(link):
    parsed = urlparse(link)
    return (parsed.netloc == "www.instagram.com" and "/p/" not in parsed.path) or \
           (parsed.netloc == "www.facebook.com" and "/posts/" not in parsed.path)

# Function to save links in an Excel file with coloring
def save_to_excel(links, filename="social_media_links.xlsx", platform="Instagram"):
    # Create a new Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"{platform} Links"

    # Add column headers
    sheet["A1"] = f"{platform} Profile Link"
    
    # Coloring the links
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green color
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow color

    # Add links with appropriate coloring
    for row_idx, link in enumerate(links, start=2):  # Starting from row 2
        sheet.cell(row=row_idx, column=1, value=link)

        # Color the link green if it contains "instagram.com"
        if platform == "Instagram" and "instagram.com" in link:
            sheet.cell(row=row_idx, column=1).fill = green_fill

        # Color the link yellow if it contains "facebook.com"
        if platform == "Facebook" and "facebook.com" in link:
            sheet.cell(row=row_idx, column=1).fill = yellow_fill

    # Save the file
    workbook.save(filename)
    print(f"{platform} Links saved to {filename}")

# Handling reCAPTCHA with random wait time
def handle_recaptcha_with_delay(driver, timeout=60):
    print("Checking for reCAPTCHA...")
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            if "recaptcha" in driver.page_source.lower():
                play_alert()
                wait_time = random.randint(5, 10)
                print(f"reCAPTCHA detected. Waiting for {wait_time} seconds...")
                time.sleep(wait_time)
            else:
                print("reCAPTCHA solved. Proceeding...")
                return True
        except Exception as e:
            logging.error(f"Error during reCAPTCHA check: {e}")
            return False
    print("Timeout reached while waiting for reCAPTCHA.")
    return False

def play_alert():
    try:
        pygame.mixer.init()  # Initialize pygame library for audio playback
        pygame.mixer.music.load("alert.mp3")  # Load the audio file (ensure the file name is correct)
        pygame.mixer.music.play()  # Play the audio
        while pygame.mixer.music.get_busy():  # Wait until the audio finishes
            pass
    except Exception as e:
        print(f"Error playing alert: {e}")

# Random delay between pages
def wait_between_pages(min_delay=5, max_delay=10):
    delay = random.uniform(min_delay, max_delay)
    print(f"Waiting for {delay:.2f} seconds before proceeding...")
    time.sleep(delay)

# Main logic for collecting links
def scrape_social_media_profiles(keywords, location, max_pages=20):
    driver = setup_driver()
    instagram_links = set()
    facebook_links = set()

    try:
        queries = generate_queries(keywords, location)
        for query in queries:
            print(f"\nSearching for: {query}")
            for page in range(max_pages):
                print(f"Scraping page {page + 1} for query '{query}'...")
                start = page * 10
                google_search_url = f"https://www.google.com/search?q={query}&start={start}"

                # Load the page with multiple attempts
                try:
                    driver.get(google_search_url)
                except Exception as e:
                    logging.error(f"Error loading {google_search_url}: {e}")
                    continue

                # Handle reCAPTCHA
                if not handle_recaptcha_with_delay(driver):
                    print("Skipping page due to unresolved reCAPTCHA.")
                    continue

                # Parse the page results
                page_source = driver.page_source
                links = parse_results(page_source)
                if not links:
                    print(f"No more results for query '{query}'.")
                    break

                for link in links:
                    if "instagram.com" in link:
                        instagram_links.add(link)
                    elif "facebook.com" in link:
                        facebook_links.add(link)

                print(f"Found {len(links)} links. Total Instagram: {len(instagram_links)}, Total Facebook: {len(facebook_links)}")
                wait_between_pages()  # Delay between pages

            print("Taking a longer break between queries...")
            wait_between_pages(30, 60)  # Long delay between queries

    except Exception as e:
        logging.error(f"Critical error: {e}")
    finally:
        driver.quit()

    # Save the links in separate Excel files
    save_to_excel(list(instagram_links), filename="instagram_links.xlsx", platform="Instagram")
    save_to_excel(list(facebook_links), filename="facebook_links.xlsx", platform="Facebook")

# Run the script
if __name__ == "__main__":
    keywords = ["your-keyword"]
    location = "your-location"
    scrape_social_media_profiles(keywords, location, max_pages=25) # Numbers of page in google search engine